import math
import pprint

from openpyxl import load_workbook, Workbook

FILENAME = "iat_se_2018.xlsx"
NAME_COL = "B"
CLASS_COL = "D"
SESSIONS = 2
CHOICES = 6
TITLE_ROW = "11"
FIRST_ROW = 13
FIRST_CHOICE_COL = "G"
BASE_COUNT_TEMPLATE = {}
CLASS_SIZES = {
    'Song Writing': 20,
    'Dance': 20,
    'Robotics': 20,
    'Entrepreneurship': 20,
    'Architecture Presentation': 20,
    'Drone Piloting': 20,
}

def load_students():
    workbook = load_workbook(FILENAME)
    signups = []
    classes = {}
    no_class_count = 0
    for school in workbook.get_sheet_names():
        cur_row = FIRST_ROW
        finished = False
        sheet = workbook[school]
        while not finished:
            cur_row_str = str(cur_row)
            new_student = {
                "name": sheet[NAME_COL + cur_row_str].value.strip(),
                "school": school,
                "class": str(sheet[CLASS_COL + cur_row_str].value).strip(),
                "choices": [],
                "assigned": []
            }
            cur_row += 1
            for i in range(CHOICES):
                new_col = chr(ord(FIRST_CHOICE_COL) + i)
                if sheet[new_col + TITLE_ROW].value.strip() not in BASE_COUNT_TEMPLATE:
                    BASE_COUNT_TEMPLATE[sheet[new_col + TITLE_ROW].value.strip()] = 0
                if sheet[new_col + cur_row_str].value:
                    class_selection = sheet[new_col + TITLE_ROW].value.replace("\n", " ").strip()
                    new_student['choices'].append(class_selection)
                    if class_selection not in classes:
                        classes[class_selection] = 0
                    classes[class_selection] += 1
            if new_student['choices'] == []:
                no_class_count += 1
            else:
                signups.append(new_student)
            if sheet[NAME_COL[0] + str(cur_row + 1)].value is None:
                finished = True
    print "no class count: " + str(no_class_count)
    return signups, classes

def get_capacity(students, classes):
    total_capacity = 0
    for course in classes:
        total_capacity += classes[course]
    return (len(students) * 1.0) / (total_capacity * 1.0)


def filter_demand(demand):
    new_demand = sorted(demand.items(), key=lambda(k,v): v)
    return [x for x, y in new_demand]

def sort_students(signups, demand, capacity):
    assigned_list = []
    session_counts = []
    incomplete_signups = []
    for i in range(SESSIONS):
        session_counts.append({})
    for signup in signups:
        choice = 0
        for j in range(SESSIONS):
            for selection in demand:
                if selection not in session_counts[j]:
                    session_counts[j] = BASE_COUNT_TEMPLATE.copy()
                if selection in signup["choices"] and session_counts[j][selection] < math.floor(CLASS_SIZES[selection] * capacity):
                    signup["session" + str(j + 1)] = selection
                    signup["assigned"].append(selection)
                    signup["choices"].remove(selection)
                    session_counts[j][selection] += 1
                    choice += 1
                    break
        if choice < SESSIONS:
            incomplete_signups.append(signup)
        else:
            assigned_list.append(signup)
    filtered_demand = []
    for i in range(SESSIONS):
        filtered_demand.append(filter_demand(session_counts[i]))
    for signup in incomplete_signups:
        for j in range(SESSIONS):
            if ("session" + str(j + 1)) not in signup:
                for choice in filtered_demand[j]:
                    if choice in signup["choices"] and session_counts[j][choice] < CLASS_SIZES[choice]:
                        signup["session" + str(j + 1)] = choice
                        signup["assigned"].append(choice)
                        signup["choices"].remove(choice)
                        session_counts[j][choice] += 1
                        break
        assigned_list.append(signup)
    pprint.pprint(session_counts)
    return assigned_list

def export_list(students, filename):
    students = sorted(students, key=lambda(d): d['name'])
    wb = Workbook()
    cur_row = 1
    cur_sheet = wb.active
    cur_sheet["A1"] = "Name"
    cur_sheet["B1"] = "School"
    cur_sheet["C1"] = "Class"
    cur_sheet["D1"] = "Session 1"
    cur_sheet["E1"] = "Session 2"
    cur_row += 1
    for student in students:
        cur_sheet["A" + str(cur_row)] = student["name"]
        cur_sheet["B" + str(cur_row)] = student["school"]
        cur_sheet["C" + str(cur_row)] = student["class"]
        cur_sheet["D" + str(cur_row)] = student["session1"]
        cur_sheet["E" + str(cur_row)] = student["session2"]
#        cur_sheet["F" + str(cur_row)] = student["session3"]
        cur_row += 1
    wb.save(filename)


if __name__ == "__main__":
    students, counts = load_students()
    capacity = get_capacity(students, CLASS_SIZES)
    demand = filter_demand(counts)
    print demand
    final_list = sort_students(students, demand, capacity)
    export_list(final_list, "temp.xlsx")

