import pprint
import os

from openpyxl import load_workbook, Workbook

SHEETNAME = "Sheet1"
NAME_COL = "B"
SCHOOL_COL = "C"
CLASS_COL = "D"
SESSIONS = 1
CHOICES = 9
TITLE_ROW = "1"
FIRST_ROW = 2
FIRST_CHOICE_COL = "E"
CLASS_SIZES = {
    'Dance': 30,
    'Digital Illustration': 25,
    'Entrepreneurship': 30,
    'Manga Illustration': 25,
    'Moble Game Building': 30,
    'Photography': 25,
    'Public Speaking': 30,
    'Robotics': 30,
    'Songwriting': 30
}
							
BASE_COUNT_TEMPLATE = {x: 0 for x in CLASS_SIZES}

def load_students():
    filenames = [filename for filename in os.listdir() if 'xlsx' in filename]
    signups = []
    classes = BASE_COUNT_TEMPLATE.copy()
    no_class_students = []

    for filename in filenames:
        workbook = load_workbook(filename)
        sheet = workbook[SHEETNAME]
        finished = False
        cur_row = FIRST_ROW
        while not finished:
            cur_row_str = str(cur_row)
            new_student = {
                "name": sheet[NAME_COL + cur_row_str].value.strip(),
                "school": sheet[SCHOOL_COL + cur_row_str].value.strip(),
                "class": sheet[CLASS_COL + cur_row_str].value,
                "choices": [],
                "assigned": []
            }
            cur_row += 1
            for i in range(CHOICES):
                new_col = chr(ord(FIRST_CHOICE_COL) + i)
                if sheet[new_col + cur_row_str].value:
                    class_selection = sheet[new_col + TITLE_ROW].value.replace("\n", " ")
                    new_student['choices'].append(class_selection)
                    if class_selection in classes:
                        classes[class_selection] += 1
            if new_student['choices'] == []:
                no_class_students.append(new_student)
            else:
                signups.append(new_student)
            if sheet[NAME_COL[0] + str(cur_row + 1)].value is None:
                finished = True
    print("no class count: " + str(len(no_class_students)))
    print(no_class_students)
    return signups, classes


def sort_demand(demand):
    new_demand = sorted(demand.items(), key=lambda demand_tuple: float(demand_tuple[1])/   float(CLASS_SIZES[demand_tuple[0]]))
    return [x for x, y in new_demand]

def sort_students(signups):
    assigned_list = []
    session_counts = []
    incomplete_signups = []
    for i in range(SESSIONS):
        session_counts.append(BASE_COUNT_TEMPLATE.copy())
    for signup in signups:
        choice = 0
        for j in range(SESSIONS):
            demand = sort_demand(session_counts[j])
            for selection in demand:
                if selection in signup["choices"] and session_counts[j][selection] < CLASS_SIZES[selection]:
                    signup["session" + str(j + 1)] = selection
                    signup["assigned"].append(selection)
                    signup["choices"].remove(selection)
                    session_counts[j][selection] += 1
                    choice += 1
                    break
        if choice < 3:
            incomplete_signups.append(signup)
        else:
            assigned_list.append(signup)
    for signup in incomplete_signups:
        for j in range(SESSIONS):
            demand = sort_demand(session_counts[j])
            if ("session" + str(j + 1)) not in signup:
                for choice in demand:
                    if choice not in signup["assigned"]:
                        signup["session" + str(j + 1)] = choice
                        signup["assigned"].append(choice)
                        session_counts[j][choice] += 1
                        break
        assigned_list.append(signup)
    pprint.pprint(session_counts)
    return assigned_list

def export_list(students, filename):
    students = sorted(students, key=lambda d: d['name'])
    wb = Workbook()
    cur_row = 1
    cur_sheet = wb.active
    cur_sheet["A1"] = "Name"
    cur_sheet["B1"] = "School"
    cur_sheet["C1"] = "Class"
    cur_sheet["D1"] = "Session 1"
    #cur_sheet["E1"] = "Session 2"
    #cur_sheet["F1"] = "Session 3"
    cur_row += 1
    for student in students:
        cur_sheet["A" + str(cur_row)] = student["name"]
        cur_sheet["B" + str(cur_row)] = student["school"]
        cur_sheet["C" + str(cur_row)] = student["class"]
        cur_sheet["D" + str(cur_row)] = student["session1"]
        #cur_sheet["E" + str(cur_row)] = student["session2"]
        #cur_sheet["F" + str(cur_row)] = student["session3"]
        cur_row += 1
    wb.save(filename)


if __name__ == "__main__":
    students, counts = load_students()
    final_list = sort_students(students)
    export_list(final_list, "deepdive.xlsx")

