import pprint

from openpyxl import load_workbook, Workbook

FILENAME = "iat_central_deepdive_signups.xlsx"
SHEETNAME = "Deep Dive Data"
NAME_COL = "D"
SCHOOL_COL = "C"
CLASS_COL = "F"
SESSIONS = 1
CHOICES = 13
TITLE_ROW = "1"
FIRST_ROW = 2
FIRST_CHOICE_COL = "I"
MAX_SIZE = 20
BASE_COUNT_TEMPLATE = {}

def load_students():
    workbook = load_workbook(FILENAME)
    signups = []
    classes = {}

    sheet = workbook[SHEETNAME]
    finished = False
    cur_row = FIRST_ROW
    while not finished:
        cur_row_str = str(cur_row)
        new_student = {
            "name": sheet[NAME_COL + cur_row_str].value.strip(),
            "school": sheet[SCHOOL_COL + cur_row_str].value.strip(),
            "choices": [],
            "assigned": []
        }
        cur_row += 1
        for i in range(CHOICES):
            new_col = chr(ord(FIRST_CHOICE_COL) + i)
            if sheet[new_col + TITLE_ROW].value not in BASE_COUNT_TEMPLATE:
                BASE_COUNT_TEMPLATE[sheet[new_col + TITLE_ROW].value] = 0
            if sheet[new_col + cur_row_str].value:
                class_selection = sheet[new_col + TITLE_ROW].value.replace("\n", " ")
                new_student['choices'].append(class_selection)
                if class_selection not in classes:
                    classes[class_selection] = 0
                classes[class_selection] += 1

        signups.append(new_student)
        if sheet[NAME_COL[0] + str(cur_row + 1)].value is None:
            finished = True
    return signups, classes


def filter_demand(demand):
    new_demand = sorted(demand.items(), key=lambda(k,v): v)
    return [x for x, y in new_demand]

def sort_students(signups, demand):
    assigned_list = []
    session_counts = []
    incomplete_signups = []
    for i in range(SESSIONS):
        session_counts.append({})
    for signup in signups:
        choice = 0
        for j in range(SESSIONS):
            for selection in demand:
                if selection not in session_counts[j]:
                    session_counts[j] = BASE_COUNT_TEMPLATE.copy()
                if selection in signup["choices"] and session_counts[j][selection] < MAX_SIZE:
                    signup["session" + str(j + 1)] = selection
                    signup["assigned"].append(selection)
                    signup["choices"].remove(selection)
                    session_counts[j][selection] += 1
                    choice += 1
                    break
        if choice < MAX_SIZE:
            incomplete_signups.append(signup)
        else:
            assigned_list.append(signup)
    filtered_demand = []
    for i in range(SESSIONS):
        filtered_demand.append(filter_demand(session_counts[i]))
    for signup in incomplete_signups:
        for j in range(SESSIONS):
            if ("session" + str(j + 1)) not in signup:
                for choice in filtered_demand[j]:
                    if choice not in signup["assigned"] and session_counts[j][choice] < MAX_SIZE:
                        signup["session" + str(j + 1)] = choice
                        signup["assigned"].append(choice)
                        session_counts[j][choice] += 1
                        break
        assigned_list.append(signup)
    print session_counts
    return assigned_list

def export_list(students, filename):
    students = sorted(students, key=lambda(d): d['name'])
    wb = Workbook()
    cur_row = 1
    cur_sheet = wb.active
    cur_sheet["A1"] = "Name"
    cur_sheet["B1"] = "School"
    cur_sheet["C1"] = "Class"
    cur_row += 1
    for student in students:
        cur_sheet["A" + str(cur_row)] = student["name"]
        cur_sheet["B" + str(cur_row)] = student["school"]
        cur_sheet["C" + str(cur_row)] = student["session1"]
        cur_row += 1
    wb.save(filename)


if __name__ == "__main__":
    students, counts = load_students()
    demand = filter_demand(counts)
    final_list = sort_students(students, demand)
    export_list(final_list, "temp.xlsx")

