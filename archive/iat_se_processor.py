from openpyxl import load_workbook, Workbook
import operator

wb = load_workbook('signups.xlsx')
sheets = wb.get_sheet_names()

students = []
classes = {}

num_sessions = 2

#read the data in
for s in sheets:
    sheet = wb[s]
    school = sheet["C4"].value
    row = 13
    finished = False
    while not finished:
        string_row = str(row)
        new_student = {}
        new_student['name'] = sheet["B" + string_row].value
        if not new_student['name']:
            finished = True
            continue
        new_student['school'] = school
        new_student['class'] = sheet["D" + string_row].value
        new_student['choices'] = []
        new_student['assigned'] = []
        for i in range(5):
            new_col = chr(ord('G') + i)
            if sheet[new_col + string_row].value:
                class_selection = sheet[new_col + "11"].value.replace("\n", " ")
                new_student['choices'].append(class_selection)
                if class_selection not in classes:
                    classes[class_selection] = 0
                classes[class_selection] += 1
        students.append(new_student)
        row += 1

#actual sorting
class_list = []
session_counts = [{},{}]

less = []

sorted_classes = sorted(classes.items(), key=operator.itemgetter(1))
sorted_classes = [x[0] for x in sorted_classes]

#sorting
for student in students:
    choice = 0
    for i in range(num_sessions):
        for class_choice in sorted_classes:
            if class_choice not in session_counts[i]:
                session_counts[i][class_choice] = 0
            if class_choice in student['choices'] and session_counts[i][class_choice] < 20 and ('session' + str(i + 1)) not in student:
                student['session' + str(i + 1)] = class_choice
                student['choices'].remove(class_choice)
                student['assigned'].append(class_choice)
                session_counts[i][class_choice] += 1
                choice += 1
                break
    if choice < 2:
        less.append(student)
    else:
        del student['choices']
        class_list.append(student)

sorted_classes = []
for session in session_counts:
    temp_sesssion_count = sorted(session.items(), key=operator.itemgetter(1))
    sorted_classes.append([x[0] for x in temp_sesssion_count])

for student in less:
    for i in range(num_sessions):
        if ('session' + str(i + 1)) not in student:
            for class_choice in sorted_classes[i]:
                if class_choice not in student['assigned'] and session_counts[i][class_choice] < 20 and ('session' + str(i + 1)) not in student:
                    student['session' + str(i + 1)] = class_choice
                    student['assigned'].append(class_choice)
                    session_counts[i][class_choice] += 1
                    break
    class_list.append(student)

print session_counts

wb = Workbook()
next_row = {}
for student in class_list:
    for i in range(num_sessions):
        if 'session' + str(i+1) in student:
            key = student['session' + str(i+1)]
            session_key = " Session " + str(i+1)
            if (key + session_key) not in wb.get_sheet_names():
                next_row[key + session_key] = 4
                wb.create_sheet(key + session_key)
                wb[key + session_key]["A1"] = key
                wb[key + session_key]["A2"] = session_key
                wb[key + session_key]["A3"] = "Name"
                wb[key + session_key]["B3"] = "School"
                wb[key + session_key]["C3"] = "Class"
            wb[key + session_key]["A" + str(next_row[key + session_key])] = student['name']
            wb[key + session_key]["B" + str(next_row[key + session_key])] = student['school']
            wb[key + session_key]["C" + str(next_row[key + session_key])] = student['class']
            next_row[key + session_key] += 1
        else:
            print student


wb.save("class_list.xlsx")
#print class_list
#print session_counts
#print less

