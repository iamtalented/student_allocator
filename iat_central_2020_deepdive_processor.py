import pprint
import os

from openpyxl import load_workbook, Workbook

SHEETNAME = "Sheet1"

# Vars for form excel
FORM_FILENAME = 'form.xlsx'
FORM_COL_STUDENT_ID = 'F'
FORM_COL_FIRST_CHOICE = 'G'
FORM_COL_SECOND_CHOICE = 'H'
FORM_ROW_FIRST_ROW = 2

# Vars for allocation excel


# Vars for programme

CLASS_SIZES = {
    'Robotics': 0,
    'Coffee Making': 1,
    'Game Design': 0,
}

BASE_COUNT_TEMPLATE = {x: 0 for x in CLASS_SIZES}

def load_selections():
    workbook = load_workbook(FORM_FILENAME)
    sheet = workbook[SHEETNAME]
    selections = {}
    duplicate_entries = []
    punk_entries = []
    finished = False
    cur_row = FORM_ROW_FIRST_ROW
    cur_row_str = str(cur_row)
    while not finished:
        cur_student_id = sheet[FORM_COL_STUDENT_ID + cur_row_str].value.strip()
        #if there is duplicate entry of same student ID, we'll take the latest one, but record down the duplicate
        if cur_student_id in selections:
            if cur_student_id not in duplicate_entries:
                duplicate_entries.append(cur_student_id)
        cur_first_choice = sheet[FORM_COL_FIRST_CHOICE + cur_row_str].value.strip()
        cur_second_choice = sheet[FORM_COL_SECOND_CHOICE + cur_row_str].value.strip()
        #record down punk entries where it's identitical for first and second choice
        if cur_first_choice == cur_second_choice:
            punk_entries.append(cur_student_id)
        selections[cur_student_id] = {
            'first_choice': cur_first_choice,
            'second_choice': cur_second_choice
        }

        #preparation for next loading cycle
        cur_row += 1
        cur_row_str = str(cur_row)
        if sheet[FORM_COL_STUDENT_ID + cur_row_str].value == None:
            finished = True
    return selections, duplicate_entries, punk_entries

"""
load data from the class master list
    - profile mapping from student id
    - session related data if students are restricted to their 3 sessions
"""
def load_more_data(selections):
    for selection in selections:
        selections[selection]['name'] = "a"
        selections[selection]['school'] = "a"
        selections[selection]['class'] = "a"
    return selections

def allocate_class(selections):
    num_choices = 2
    choice_keys = ['first_choice', 'second_choice']
    selections_holder = [selections, {}]
    session_count = BASE_COUNT_TEMPLATE.copy()
    class_list = []
    no_class = {}
    for i in range(num_choices):
        for student_id in selections_holder[i]:
            selection = selections[student_id][choice_keys[i]]
            if session_count[selection] < CLASS_SIZES[selection]:
                # allocate student to class
                selections[student_id]["student_id"] = student_id
                selections[student_id]['final_choice'] = selection
                class_list.append(selections[student_id])
                session_count[selection] += 1
            elif i + 1 < num_choices:
                # pass student to next batch of sorting
                selections_holder[i + 1][student_id] = True
            else:
                no_class[student_id] = selections[student_id]
    return class_list, no_class

def export_list(students, filename):
    print students
    students = sorted(students, key=lambda d: d['name'])
    wb = Workbook()
    cur_row = 1
    cur_sheet = wb.active
    cur_sheet["A1"] = "Name"
    cur_sheet["B1"] = "School"
    cur_sheet["C1"] = "Class"
    cur_sheet["D1"] = "Student ID"
    cur_sheet["E1"] = "Final Session"
    cur_row += 1
    for student in students:
        cur_sheet["A" + str(cur_row)] = student["name"]
        cur_sheet["B" + str(cur_row)] = student["school"]
        cur_sheet["C" + str(cur_row)] = student["class"]
        cur_sheet["D" + str(cur_row)] = student["student_id"]
        cur_sheet["E" + str(cur_row)] = student["final_choice"]
        cur_row += 1
    wb.save(filename)


if __name__ == "__main__":
    selections, duplicate_entries, punk_entries = load_selections()
    selections = load_more_data(selections)
    final_list, no_class = allocate_class(selections)
    print("No Class: " + str(no_class))
    export_list(final_list, "deepdive.xlsx")

