import pprint

from openpyxl import load_workbook, Workbook

FILENAME = "collaborate_signups.xlsx"
FIRST_ROW = 2
NAME_COL = ["D", "E"]
SCHOOL_COL = "H"
EMAIL_COL = "B"
PHONE_COL = "K"
CHOICE_COL = "L"
CONSENT_COL = "P"
SESSIONS = 3
MAX_SIZE = 7
MIN_TO_START = 12
BOOKS = ["Accounting","Banking (Consumer / Investment)","Computing / Technology","Education","Engineering","Government (Central Admin)","Government (Economic)","Government (HR)","Government (Infrastructure)","Government (Social)","Healthcare (Allied)","Healthcare (Medical)","Journalism","Legal","Linguistics","Management Consulting","Social Enterprise","Social Work"]

def load_students():
    workbook = load_workbook(FILENAME)
    signups = []
    emails = []
    choices = {}
    for sheet_name in workbook.get_sheet_names():
        sheet = workbook[sheet_name]
        finished = False
        cur_row = FIRST_ROW
        while not finished:
            cur_row_str = str(cur_row)
            new_signup = {
                "name": sheet[NAME_COL[0] + cur_row_str].value + " " +
                        sheet[NAME_COL[1] + cur_row_str].value,
                "school": sheet[SCHOOL_COL + cur_row_str].value,
                "email": sheet[EMAIL_COL + cur_row_str].value,
                "phone": sheet[PHONE_COL + cur_row_str].value,
                "choices": [x for x in BOOKS if sheet[CHOICE_COL + cur_row_str].value.find(x) != -1],
                "consent": sheet[CONSENT_COL + cur_row_str].value,
                "assigned": []
            }
            cur_row += 1
            if sheet[NAME_COL[0] + str(cur_row + 1)].value is None:
                finished = True
            if new_signup["email"] in emails:
                print("inhere")
                continue
            emails.append(new_signup["email"])
            for choice in new_signup["choices"]:
                if choice not in choices:
                    choices[choice] = 0
                choices[choice] += 1
            signups.append(new_signup)
    return signups, choices

def filter_demand(demand, cutoff):
    books = BOOKS
    new_demand = sorted([(k,v) for k,v in demand.items() if k in books], key=lambda demand_tuple: demand_tuple[1])
    return [x for x, y in new_demand]

def sort_students(signups, filtered_demand):
    assigned_list = []
    session_counts = []
    incomplete_signups = []
    for i in range(SESSIONS):
        session_counts.append({})
    for signup in signups:
        choice = 0
        for j in range(SESSIONS):
            for selection in filtered_demand:
                if selection not in session_counts[j]:
                    session_counts[j][selection] = 0
                if selection in signup["choices"] and session_counts[j][selection] < MAX_SIZE:
                    signup["session" + str(j + 1)] = selection
                    signup["assigned"].append(selection)
                    signup["choices"].remove(selection)
                    session_counts[j][selection] += 1
                    choice += 1
                    break
        if choice < MAX_SIZE:
            incomplete_signups.append(signup)
        else:
            assigned_list.append(signup)
    filtered_demand = []
    for i in range(SESSIONS):
        filtered_demand.append(filter_demand(session_counts[i], -1))
    for signup in incomplete_signups:
        for j in range(SESSIONS):
            if ("session" + str(j + 1)) not in signup:
                for choice in filtered_demand[j]:
                    if choice not in signup["assigned"] and session_counts[j][choice] < MAX_SIZE:
                        signup["session" + str(j + 1)] = choice
                        signup["assigned"].append(choice)
                        session_counts[j][choice] += 1
                        break
        assigned_list.append(signup)
    return assigned_list

def export_books(readers, filename):
    readers = sorted(readers, key=lambda d: d['name'])
    wb = Workbook()
    cur_row = 1
    cur_sheet = wb.active
    cur_sheet["A1"] = "Name"
    cur_sheet["B1"] = "Book 1"
    cur_sheet["C1"] = "Book 2"
    cur_sheet["D1"] = "Book 3"
    cur_row += 1
    for reader in readers:
        cur_sheet["A" + str(cur_row)] = reader["name"]
        cur_sheet["B" + str(cur_row)] = reader["session1"]
        cur_sheet["C" + str(cur_row)] = reader["session2"]
        cur_sheet["D" + str(cur_row)] = reader["session3"]
        cur_row += 1
    wb.save(filename)


if __name__ == "__main__":
    students, counts = load_students()
    filtered_demand = filter_demand(counts, MIN_TO_START)
    assigned_list = sort_students(students, filtered_demand)
    for signup in assigned_list:
        if len(signup["assigned"]) < SESSIONS:
            print(signup)
    export_books(assigned_list, "collaborate_particpants.xlsx")

