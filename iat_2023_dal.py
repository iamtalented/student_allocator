import os
from pprint import pprint
import random
import string

from openpyxl import load_workbook, Workbook

DATA_DIR = "./data/"
EXPORT_FILENAME = "Class List.xlsx"
EXPORT_MAINLIST_SHEETNAME = "Main List"
FIRST_CHOICE_COL = "E"
FIRST_ROW = 6
NAME_COL = "B"
NUM_CLASSES = 10
OUTPUT_DIR = "./output/"
SCHOOL_COL = "D"
SESSIONS = 3
SHEETNAME = "Reg List - IAT2024"
TITLE_ROW = "4"

def load_students_tasters():
    filenames = [filename for filename in os.listdir(DATA_DIR) if 'xlsx' in filename]
    students = []
    for filename in filenames:
        students.extend(load_students_from_excel_tasters(DATA_DIR+filename))
    return students

def load_students_from_excel_tasters(filename):
    workbook = load_workbook(filename)
    
    sheet = workbook[SHEETNAME]
    signups = []
    cur_row = FIRST_ROW
    finished = False
    generated_ids = set()
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        while not finished:
            cur_row_str = str(cur_row)
            if sheet[NAME_COL[0] + cur_row_str].value is None:
                finished = True
                break
            new_id = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
            while new_id in generated_ids:
                new_id = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
            generated_ids.add(new_id)
            new_student = {
                "name": sheet[NAME_COL + cur_row_str].value.strip(),
                "school": sheet[SCHOOL_COL + cur_row_str].value.strip(),
                "id": new_id,
                "choices": [],
                "assigned": []
            }
            cur_row += 1
            for i in range(NUM_CLASSES):
                new_col = chr(ord(FIRST_CHOICE_COL) + i)
                if sheet[new_col + cur_row_str].value:
                    class_selection = sheet[new_col + TITLE_ROW].value.replace("\n", " ")
                    new_student['choices'].append(class_selection)
            if new_student['choices'] == []:
                no_class_count += 1
            else:
                signups.append(new_student)

    return signups

def export_student_tasters_to_excel(students):
    students = sorted(students, key=lambda student: student["school"])
    wb = Workbook()
    wb.create_sheet(EXPORT_MAINLIST_SHEETNAME)
    write_to_student_list_to_sheet(wb[EXPORT_MAINLIST_SHEETNAME], students)

    class_list = generate_class_lists(students)
    for course in class_list:
        course = str(course)
        wb.create_sheet(course)
        write_class_list_to_sheet(wb[course], class_list[course], course)
    
    wb.save(OUTPUT_DIR + EXPORT_FILENAME)
    

def generate_class_lists(students):
    class_list = {}
    for student in students:
        for i in range(SESSIONS):
            cur_session = "session" + str(i+1)
            if student[cur_session] not in class_list:
                class_list[student[cur_session]] = {
                    "session1": [],
                    "session2": [],
                    "session3": []
                }
            class_list[student[cur_session]][cur_session].append((student["name"], student["school"], student["id"]))
    return class_list

def write_to_student_list_to_sheet(sheet, students):
    sheet["A1"] = "Name"
    sheet["B1"] = "School"
    sheet["C1"] = "ID"
    sheet["D1"] = "Session 1"
    sheet["E1"] = "Session 2"
    sheet["F1"] = "Session 3"
    cur_row = 2
    for student in students:
        cur_row_str = str(cur_row)
        sheet["A" + cur_row_str] = student["name"]
        sheet["B" + cur_row_str] = student["school"]
        sheet["C" + cur_row_str] = student["id"]
        sheet["D" + cur_row_str] = student["session1"]
        sheet["E" + cur_row_str] = student["session2"]
        sheet["F" + cur_row_str] = student["session3"]
        cur_row += 1


def write_class_list_to_sheet(sheet, class_list, course_name):
    sheet["A1"] = course_name
    cur_row = 2
    for i in range(SESSIONS):
        cur_session = "session" + str(i+1)
        sheet["A" + str(cur_row)] = "SESSION " + str(i+1)
        cur_row += 1
        for student, school, id in class_list[cur_session]:
            sheet["A" + str(cur_row)] = student
            sheet["B" + str(cur_row)] = school
            sheet["C" + str(cur_row)] = id
            cur_row += 1
        cur_row += 1
    



